# -*- coding: utf-8 -*-
"""Salidas del radar (v1.3): Excel de detecciones, dashboard HTML (GitHub Pages),
informe de los últimos N días, fichero de alertas para el Issue y — nuevo —
versiones HTML de informe y alertas con enlaces clicables para el correo."""
from datetime import timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .util import DATA, DOCS, ahora_utc

AZUL = "1F3864"


def _hoja(ws, titulo, columnas, anchos):
    ws.append([])
    for i, (h, w) in enumerate(zip(columnas, anchos), 1):
        c = ws.cell(1, i, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fila_det(d):
    s = d.get("senales", {})
    return [
        d.get("fecha_deteccion"), d.get("score"), d.get("fuente"),
        d.get("pais", ""), d.get("entidad_censo") or d.get("organo", ""),
        d.get("titulo", ""), d.get("importe_eur"),
        "Sí" if d.get("es_redaccion_proyecto") else "",
        ", ".join(s.get("sin_zanja", [])),
        d.get("plazo_presentacion", ""), d.get("estado", ""),
        ", ".join(d.get("cpv", [])[:5]), d.get("enlace", ""),
    ]


COLS = ["Detectado", "Score", "Fuente", "País", "Entidad / órgano", "Objeto",
        "Importe (€)", "Redacción proyecto", "Señales sin zanja",
        "Plazo presentación", "Estado", "CPV", "Enlace"]
ANCHOS = [11, 7, 8, 9, 28, 55, 13, 10, 22, 14, 14, 18, 45]


def generar_excel(historico, nuevas, ajustes):
    """Genera el Excel Histórico (repositorio analítico) en docs/, servido por
    GitHub Pages en una URL fija. Diseño según el Canalis Brand Book."""
    from . import excel_historico
    excel_historico.generar(historico, {"interno": EXCEL_INTERNO_URL, "panel": PANEL_URL})


# ------------------------------------------------ correo HTML (estilo v2.0)
# Gramatica visual tomada del Canalis Brand Book:
# fondo hueso #EBECE6, rojo #E03C32 como unico acento, Helvetica Neue,
# tarjetas planas con cifra gigante, filas tipo indice, mucho aire.
# Imagenes: subir a la carpeta docs/ con estos nombres exactos.
LOGO_URL = "https://ibarrez.github.io/canalis-tender/logo.png"        # logotipo canalis en NEGRO, PNG fondo transparente
FOTO_URL = "https://ibarrez.github.io/canalis-tender/cabecera.jpg"
EXCEL_HISTORICO_URL = "https://ibarrez.github.io/canalis-tender/Radar_Historico_ES-PT.xlsx"
PANEL_URL = "https://ibarrez.github.io/canalis-tender/"    # foto robot estudio blanco
EXCEL_INTERNO_URL = "https://canalisst-my.sharepoint.com/:x:/g/personal/jacobo_rodriguez_grupocanalis_com/IQBBLqwTIrSmT5fiVcrQ-Jc0AXzAml8M9y6b6MABa_ffujk?e=Ah2fYE"

NEGRO = "#111111"
GRIS = "#6E6E69"
GRISCLARO = "#B6B7B0"
HUESO = "#EBECE6"
LINEA = "#DADBD5"
ROJO = "#E03C32"
F = "'Helvetica Neue',Helvetica,Arial,sans-serif"


def _esc(t):
    return (str(t or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _millones(v):
    if not v:
        return ""
    if v >= 1_000_000:
        return (f"{v/1_000_000:.1f}".replace(".", ",") + "M")
    return f"{v/1000:.0f}K"


def _item_html(i, d):
    ent = _esc(d.get("entidad_censo") or d.get("organo", ""))
    score = d.get("score", 0)
    color_score = ROJO if score >= 75 else GRISCLARO
    meta = []
    if d.get("importe_eur"):
        imp = f"{d['importe_eur']:,.0f}".replace(",", ".")
        meta.append(f'<span style="color:{NEGRO};font-weight:bold">{imp} EUR</span>')
    if d.get("plazo_presentacion"):
        meta.append(f'plazo {_esc(d["plazo_presentacion"])}')
    sz = d.get("senales", {}).get("sin_zanja", [])
    if sz:
        meta.append(f'sin zanja: {_esc(", ".join(sz))}')
    if d.get("es_redaccion_proyecto"):
        meta.append(f'<span style="color:{ROJO}">redacción de proyecto</span>')
    linea_meta = "&nbsp;&nbsp;/&nbsp;&nbsp;".join(meta)
    return f"""<tr><td style="padding:20px 0;border-top:1px solid {LINEA}">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
<td width="38" valign="top" style="font-family:{F};font-size:13px;color:{GRISCLARO};padding-top:2px">{i:02d}</td>
<td valign="top" style="font-family:{F}">
  <div style="font-size:15px;font-weight:bold;color:{NEGRO};letter-spacing:-0.2px">{ent}</div>
  <div style="font-size:14px;color:{GRIS};line-height:1.5;margin:4px 0 8px">{_esc(d.get("titulo", ""))[:220]}</div>
  <div style="font-size:12px;color:{GRIS};line-height:1.6">{linea_meta}</div>
  <div style="margin-top:10px"><a href="{d.get("enlace", "#")}" style="font-family:{F};font-size:11px;letter-spacing:1.5px;color:{NEGRO};text-decoration:none;font-weight:bold">VER EXPEDIENTE <span style="color:{ROJO}">&#8594;</span></a></div>
</td>
<td width="64" valign="top" align="right" style="font-family:{F};font-size:26px;font-weight:bold;letter-spacing:-1px;color:{color_score}">{score}</td>
</tr></table></td></tr>"""


def _item_adj_html(i, d):
    ent = _esc(d.get("entidad_censo") or d.get("organo", ""))
    ganadores = d.get("adjudicatarios") or []
    gan = _esc(" + ".join(ganadores)) if ganadores else "No consta en la publicacion"
    color_gan = NEGRO if ganadores else GRISCLARO
    meta = []
    if d.get("importe_adjudicacion"):
        imp = f"{d['importe_adjudicacion']:,.0f}".replace(",", ".")
        meta.append(f'<span style="color:{NEGRO};font-weight:bold">{imp} EUR</span>')
    elif d.get("importe_eur"):
        imp = f"{d['importe_eur']:,.0f}".replace(",", ".")
        meta.append(f'{imp} EUR licitacion')
    if d.get("num_ofertas"):
        meta.append(f'{d["num_ofertas"]} ofertas')
    if d.get("fecha_adjudicacion"):
        meta.append(f'{_esc(d["fecha_adjudicacion"])}')
    linea_meta = "&nbsp;&nbsp;/&nbsp;&nbsp;".join(meta)
    return f"""<tr><td style="padding:18px 0;border-top:1px solid {LINEA}">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
<td width="38" valign="top" style="font-family:{F};font-size:13px;color:{GRISCLARO};padding-top:2px">{i:02d}</td>
<td valign="top" style="font-family:{F}">
  <div style="font-size:14px;font-weight:bold;color:{color_gan};letter-spacing:-0.2px">{gan}</div>
  <div style="font-size:12px;color:{GRIS};margin:3px 0 6px">{ent}</div>
  <div style="font-size:13px;color:{GRIS};line-height:1.5;margin:0 0 8px">{_esc(d.get("titulo", ""))[:180]}</div>
  <div style="font-size:12px;color:{GRIS}">{linea_meta}</div>
  <div style="margin-top:8px"><a href="{d.get("enlace", "#")}" style="font-family:{F};font-size:11px;letter-spacing:1.5px;color:{NEGRO};text-decoration:none;font-weight:bold">VER EXPEDIENTE <span style="color:{ROJO}">&#8594;</span></a></div>
</td></tr></table></td></tr>"""


def _cabecera_seccion(titulo, subtitulo):
    return f"""<tr><td style="padding:38px 0 16px">
<table role="presentation" cellpadding="0" cellspacing="0"><tr>
<td width="10" bgcolor="{ROJO}" style="width:10px;height:10px;font-size:0;line-height:0">&nbsp;</td>
<td style="font-family:{F};font-size:17px;font-weight:bold;color:{NEGRO};padding-left:12px;letter-spacing:-0.2px">{_esc(titulo)}</td>
</tr></table>
<div style="font-family:{F};font-size:12px;color:{GRIS};margin-top:8px">{_esc(subtitulo)}</div>
</td></tr>"""


def _kpi_html(kpis):
    celdas = ""
    for i, (label, valor) in enumerate(kpis[:3]):
        pad = "padding-right:8px" if i < len(kpis[:3]) - 1 else ""
        celdas += f"""<td width="33%" valign="top" style="{pad}">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
<td bgcolor="{HUESO}" style="padding:16px 16px 18px">
<div style="font-family:{F};font-size:12px;color:{NEGRO};line-height:1.4;min-height:34px">{_esc(label)}</div>
<div style="font-family:{F};font-size:42px;font-weight:bold;letter-spacing:-2px;color:{NEGRO};margin-top:22px">{_esc(valor)}</div>
</td></tr></table></td>"""
    return f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>{celdas}</tr></table>'


def _envoltorio_html(titulo, subtitulo, cuerpo, kpis=None):
    bloque_kpi = f'<tr><td bgcolor="#FFFFFF" style="padding:6px 36px 26px">{_kpi_html(kpis)}</td></tr>' if kpis else ""
    return f"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"></head>
<body style="margin:0;padding:0;background:{HUESO}">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" bgcolor="{HUESO}"><tr><td align="center" style="padding:30px 14px">
<table role="presentation" width="680" cellpadding="0" cellspacing="0" style="max-width:680px;width:100%">

<tr><td bgcolor="#FFFFFF" style="padding:26px 36px 22px">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>
<td><img src="{LOGO_URL}" alt="canalis" height="22" style="display:inline-block;border:0;height:22px"></td>
<td align="right" style="font-family:{F};font-size:10px;line-height:1.6;color:{NEGRO};text-align:right">Radar de licitaciones<br><span style="color:{GRIS}">España y Portugal</span></td>
</tr></table></td></tr>

<tr><td bgcolor="#FFFFFF"><img src="{FOTO_URL}" alt="" width="680" style="display:block;border:0;width:100%"></td></tr>

<tr><td bgcolor="#FFFFFF" style="padding:36px 36px 24px">
<table role="presentation" cellpadding="0" cellspacing="0"><tr>
<td width="10" bgcolor="{ROJO}" style="width:10px;height:10px;font-size:0;line-height:0">&nbsp;</td>
<td style="font-family:{F};font-size:17px;font-weight:bold;color:{NEGRO};padding-left:12px;letter-spacing:-0.2px">{_esc(titulo)}</td>
</tr></table>
<div style="font-family:{F};font-size:12px;color:{GRIS};margin-top:10px">{_esc(subtitulo)}</div>
</td></tr>

{bloque_kpi}

<tr><td bgcolor="#FFFFFF" style="padding:0 36px 10px">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0">{cuerpo}</table>
</td></tr>

<tr><td bgcolor="#FFFFFF" style="padding:22px 36px 30px">
<div style="border-top:1px solid {LINEA};padding-top:16px">
<a href="{EXCEL_HISTORICO_URL}" style="font-family:{F};font-size:11px;letter-spacing:1.5px;color:{NEGRO};text-decoration:none;font-weight:bold">EXCEL HIST&Oacute;RICO <span style="color:{ROJO}">&#8594;</span></a>&nbsp;&nbsp;&nbsp;&nbsp;<a href="{EXCEL_INTERNO_URL}" style="font-family:{F};font-size:11px;letter-spacing:1.5px;color:{NEGRO};text-decoration:none;font-weight:bold">EXCEL INTERNO <span style="color:{ROJO}">&#8594;</span></a>
<div style="font-family:{F};font-size:10px;color:{GRISCLARO};line-height:1.6;margin-top:12px">Verifica siempre las fechas e importes en la fuente oficial antes de actuar y traslada las oportunidades maduras al Excel interno<br>Desarrollado por Jacobo Ib&aacute;rrez</div>
</div></td></tr>

</table></td></tr></table></body></html>"""


ESTADOS_CERRADOS = {"ADJ", "RES", "ANUL"}  # adjudicada, resuelta/formalizada, anulada (códigos PLACSP)


def _vigente(d):
    """True si la oportunidad sigue viva: no adjudicada/resuelta y con plazo no vencido.
    Sin datos de estado o plazo se asume viva (mejor avisar de más que perder una)."""
    if str(d.get("estado", "")).strip().upper() in ESTADOS_CERRADOS:
        return False
    plazo = str(d.get("plazo_presentacion", ""))[:10]
    if len(plazo) == 10 and plazo < ahora_utc().strftime("%Y-%m-%d"):
        return False
    return True


def generar_informe(historico, ajustes):
    dias = int(ajustes.get("dias_ventana_informe", 7))
    corte = (ahora_utc() - timedelta(days=dias)).strftime("%Y-%m-%d")
    recientes = [d for d in historico.values() if d.get("fecha_deteccion", "") >= corte]
    top = sorted((d for d in recientes if _vigente(d)), key=lambda x: -x.get("score", 0))[:10]
    lineas = [f"# Informe del radar — {ahora_utc().strftime('%d/%m/%Y')}",
              f"\nDetecciones de los últimos {dias} días: **{len(recientes)}**. Top 10 por puntuación:\n"]
    if not top:
        lineas.append("_Sin detecciones relevantes en la ventana. Revisa data/estado.json por si alguna fuente falló._")
    for i, d in enumerate(top, 1):
        imp = f" · {d['importe_eur']:,.0f} €".replace(",", ".") if d.get("importe_eur") else ""
        rp = " · **REDACCIÓN DE PROYECTO**" if d.get("es_redaccion_proyecto") else ""
        sz = f" · sin zanja: {', '.join(d['senales']['sin_zanja'])}" if d.get("senales", {}).get("sin_zanja") else ""
        lineas.append(f"{i}. **[{d.get('score', 0)}]** {d.get('entidad_censo') or d.get('organo', '¿?')} — "
                      f"{d.get('titulo', '')[:180]}{imp}{rp}{sz} · [{d.get('fuente')}]({d.get('enlace', '')})")
    adjudicadas = sorted(
        [d for d in recientes
         if d.get("adjudicatarios") or str(d.get("estado", "")).strip().upper() in ("ADJ", "RES")],
        key=lambda x: -(x.get("importe_adjudicacion") or x.get("importe_eur") or 0))[:10]
    if adjudicadas:
        lineas.append("\n## Adjudicaciones de la semana\n")
        for d in adjudicadas:
            gan = " + ".join(d.get("adjudicatarios") or []) or "adjudicatario no consta"
            imp = f" por {d['importe_adjudicacion']:,.0f} EUR".replace(",", ".") if d.get("importe_adjudicacion") else ""
            of = f" ({d['num_ofertas']} ofertas)" if d.get("num_ofertas") else ""
            lineas.append(f"- **{gan}**{imp}{of}: {d.get('organo', '')}: {d.get('titulo', '')[:140]} ({d.get('enlace', '')})")
    lineas.append("\n---\n_Radar automático: verifica siempre fechas e importes en la fuente oficial antes de actuar. "
                  "Traslada las oportunidades maduras al Excel interno (ficha comercial + scoring de 13 criterios)._")
    (DATA / "informe.md").write_text("\n".join(lineas), encoding="utf-8")
    items = "".join(_item_html(i, d) for i, d in enumerate(top, 1)) or \
        f'<tr><td style="font-family:{F};font-size:13px;color:{GRIS};padding:14px 0"><i>Sin detecciones relevantes en la ventana.</i></td></tr>'
    kpis = [("Detectadas en la última semana", str(len(recientes))),
            ("Importe agregado", _millones(sum(d.get("importe_eur") or 0 for d in top)) or "0"),
            ("En fase de redacción", str(sum(1 for d in top if d.get("es_redaccion_proyecto"))))]
    cuerpo = items
    if adjudicadas:
        cuerpo += _cabecera_seccion("Adjudicaciones de la semana",
                                    "Quién ha ganado qué.")
        cuerpo += "".join(_item_adj_html(i, d) for i, d in enumerate(adjudicadas, 1))
    (DATA / "informe.html").write_text(
        _envoltorio_html(f"Informe del radar {ahora_utc().strftime('%d/%m/%Y')}",
                         "Rehabilitación y renovación de redes de agua. Tecnologías sin zanja.",
                         cuerpo, kpis), encoding="utf-8")


def generar_alertas(nuevas, ajustes):
    """Escribe alertas_nuevas.md (Issue) y alertas_nuevas.html (correo) solo si
    hay novedades ≥ umbral. Si no las hay, elimina ambos ficheros."""
    ruta_md = DATA / "alertas_nuevas.md"
    ruta_html = DATA / "alertas_nuevas.html"
    for r in (ruta_md, ruta_html):
        if r.exists():
            r.unlink()
    umbral = ajustes.get("umbral_alerta", 60)
    fuertes = sorted([d for d in nuevas if d.get("score", 0) >= umbral and _vigente(d)],
                     key=lambda x: -x["score"])
    if not fuertes:
        return 0
    lineas = [f"## {len(fuertes)} detección(es) nueva(s) con score ≥ {umbral}\n"]
    for d in fuertes[:25]:
        imp = f" · {d['importe_eur']:,.0f} €".replace(",", ".") if d.get("importe_eur") else ""
        lineas.append(f"- **[{d['score']}]** {d.get('entidad_censo') or d.get('organo', '¿?')} — "
                      f"{d.get('titulo', '')[:200]}{imp} → {d.get('enlace', '')}")
    lineas.append("\nRevisar, verificar en fuente oficial y pasar al Excel maestro si procede.")
    ruta_md.write_text("\n".join(lineas), encoding="utf-8")
    items = "".join(_item_html(i, d) for i, d in enumerate(fuertes[:25], 1))
    kpis = [("Detectadas en este barrido", str(len(fuertes))),
            ("Importe agregado", _millones(sum(d.get("importe_eur") or 0 for d in fuertes)) or "0"),
            ("En fase de redacción", str(sum(1 for d in fuertes if d.get("es_redaccion_proyecto"))))]
    ruta_html.write_text(
        _envoltorio_html("Nuevas oportunidades detectadas",
                         f"Barrido del {ahora_utc().strftime('%d/%m/%Y')}. PLACSP, TED y BASE. Score {umbral} o superior.",
                         items, kpis), encoding="utf-8")
    return len(fuertes)


def _fila_html(d):
    imp = f"{d['importe_eur']:,.0f} €".replace(",", ".") if d.get("importe_eur") else "—"
    rp = "📐" if d.get("es_redaccion_proyecto") else ""
    sz = "🟢" if d.get("senales", {}).get("sin_zanja") else ""
    ent = d.get("entidad_censo") or d.get("organo", "")
    return (f"<tr><td class='sc'>{d.get('score', 0)}</td><td>{d.get('fecha_deteccion', '')}</td>"
            f"<td>{d.get('fuente', '')}</td><td>{_esc(ent)[:60]}</td>"
            f"<td><a href='{d.get('enlace', '#')}' target='_blank'>{_esc(d.get('titulo') or '')[:150]}</a> {rp}{sz}</td>"
            f"<td class='imp'>{imp}</td></tr>")


def generar_dashboard(historico, estado, ajustes):
    DOCS.mkdir(exist_ok=True)
    todas = sorted(historico.values(), key=lambda x: (-x.get("score", 0), x.get("fecha_deteccion", "")))
    corte7 = (ahora_utc() - timedelta(days=7)).strftime("%Y-%m-%d")
    nuevas7 = [d for d in historico.values() if d.get("fecha_deteccion", "") >= corte7]
    redaccion = [d for d in todas if d.get("es_redaccion_proyecto")]
    filas_top = "\n".join(_fila_html(d) for d in todas[:50])
    filas_rp = "\n".join(_fila_html(d) for d in redaccion[:30]) or "<tr><td colspan='6'>Sin señales de redacción de proyecto todavía.</td></tr>"
    avisos = "".join(f"<li>{_esc(a)}</li>" for a in estado.get("registro", []) if "aviso" in a)
    html = f"""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Radar de Licitaciones de Agua ES-PT</title>
<style>
 body{{font-family:Arial,Helvetica,sans-serif;margin:0;background:#f5f7fa;color:#1a1a2e}}
 header{{background:#1F3864;color:#fff;padding:22px 28px}}
 header h1{{margin:0;font-size:21px}} header p{{margin:6px 0 0;font-size:13px;opacity:.85}}
 .kpis{{display:flex;gap:14px;padding:18px 28px;flex-wrap:wrap}}
 .kpi{{background:#fff;border-radius:10px;padding:14px 20px;box-shadow:0 1px 4px rgba(0,0,0,.08);min-width:150px}}
 .kpi b{{font-size:26px;color:#1F3864;display:block}} .kpi span{{font-size:12px;color:#555}}
 section{{margin:8px 28px 26px}} h2{{font-size:16px;color:#1F3864}}
 table{{width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08);font-size:13px}}
 th{{background:#2E5AA8;color:#fff;text-align:left;padding:8px 10px;font-size:12px}}
 td{{padding:7px 10px;border-top:1px solid #eef1f6;vertical-align:top}}
 td.sc{{font-weight:bold;color:#1F3864;text-align:center}} td.imp{{white-space:nowrap;text-align:right}}
 a{{color:#2E5AA8;text-decoration:none}} a:hover{{text-decoration:underline}}
 .avisos{{background:#FFF2CC;border-radius:10px;padding:10px 16px;font-size:12px}}
 footer{{padding:14px 28px 30px;font-size:11px;color:#777}}
 .leyenda{{font-size:12px;color:#555;margin:6px 0 12px}}
</style></head><body>
<header><h1>💧 Radar de Licitaciones de Agua - España y Portugal</h1>
<p>Vigilancia automática de rehabilitación y renovación de redes (tecnologías sin zanja) · Última ejecución: {estado.get('ejecutado', '—')} UTC</p></header>
<div class="kpis">
 <div class="kpi"><b>{len(historico)}</b><span>detecciones acumuladas</span></div>
 <div class="kpi"><b>{len(nuevas7)}</b><span>en los últimos 7 días</span></div>
 <div class="kpi"><b>{len(redaccion)}</b><span>señales de redacción de proyecto</span></div>
 <div class="kpi"><b>{estado.get('nuevas_alerta', 0)}</b><span>alertas nuevas (score ≥ {ajustes.get('umbral_alerta', 60)})</span></div>
</div>
<section><h2>Top 50 por puntuación</h2>
<div class="leyenda">📐 = fase de redacción de proyecto (señal temprana) · 🟢 = menciona tecnología sin zanja · el score automático mide relevancia de la señal, no sustituye al análisis comercial</div>
<table><tr><th>Score</th><th>Detectado</th><th>Fuente</th><th>Entidad / órgano</th><th>Objeto</th><th>Importe</th></tr>
{filas_top}</table></section>
<section><h2>📐 Radar de redacciones de proyecto (la señal más temprana)</h2>
<table><tr><th>Score</th><th>Detectado</th><th>Fuente</th><th>Entidad / órgano</th><th>Objeto</th><th>Importe</th></tr>
{filas_rp}</table></section>
{f'<section><h2>Avisos de fuentes</h2><div class="avisos"><ul>{avisos}</ul></div></section>' if avisos else ''}
<footer>Generado automáticamente. Verifica siempre fechas e importes en la fuente oficial (PLACSP · TED · BASE.gov.pt) antes de cualquier acción comercial.
El Excel completo está en <code>data/Radar_Detecciones.xlsx</code> del repositorio; el informe semanal en <code>data/informe.md</code>.</footer>
</body></html>"""
    (DOCS / "index.html").write_text(html, encoding="utf-8")
