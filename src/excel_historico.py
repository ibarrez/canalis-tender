# -*- coding: utf-8 -*-
"""Excel Histórico del Radar: repositorio analítico regenerado en cada barrido.

Diseño según el Canalis Brand Book: Helvetica, fondo hueso #EBECE6, negro,
rojo #E03C32 como único acento (urgencias, scores altos, marcadores de
sección). Pensado para que un usuario básico lo entienda y controle con los
filtros de cabecera, con calidad de información de nivel analítico.

Hojas: Portada (KPIs + cómo usar) · Vigentes (priorizadas por urgencia) ·
Adjudicaciones (quién gana qué) · Ranking Ganadores (top 15) ·
Análisis (por CCAA y por servicio) · Histórico por año.
"""
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .util import DOCS, ahora_utc, normalizar

# ------------------------------------------------- Canalis Brand Book
NEGRO = "111111"
GRIS = "6E6E69"
GRISCLARO = "B6B7B0"
HUESO = "EBECE6"
ROJO = "E03C32"
BLANCO = "FFFFFF"
FN = "Helvetica Neue"

F_TITULO = Font(name=FN, size=16, bold=True, color=NEGRO)
F_SUB = Font(name=FN, size=10, color=GRIS)
F_HDR = Font(name=FN, size=9, bold=True, color=NEGRO)
F_TXT = Font(name=FN, size=10, color=NEGRO)
F_TXT_GRIS = Font(name=FN, size=10, color=GRIS)
F_KPI_LBL = Font(name=FN, size=10, color=NEGRO)
F_KPI_NUM = Font(name=FN, size=34, bold=True, color=NEGRO)
F_ROJO = Font(name=FN, size=10, bold=True, color=ROJO)
F_NEGRITA = Font(name=FN, size=10, bold=True, color=NEGRO)
FILL_HUESO = PatternFill("solid", fgColor=HUESO)
FILL_ROJO = PatternFill("solid", fgColor=ROJO)
LINEA = Border(bottom=Side(style="thin", color="DADBD5"))
WRAP = Alignment(wrap_text=True, vertical="top")
CENTRO = Alignment(horizontal="center", vertical="top")
DER = Alignment(horizontal="right", vertical="top")

EUR = '#,##0 "EUR"'

NUTS2_CCAA = {
    "ES11": "Galicia", "ES12": "Asturias", "ES13": "Cantabria", "ES21": "País Vasco",
    "ES22": "Navarra", "ES23": "La Rioja", "ES24": "Aragón", "ES30": "Madrid",
    "ES41": "Castilla y León", "ES42": "Castilla-La Mancha", "ES43": "Extremadura",
    "ES51": "Cataluña", "ES52": "Comunitat Valenciana", "ES53": "Illes Balears",
    "ES61": "Andalucía", "ES62": "Región de Murcia", "ES63": "Ceuta", "ES64": "Melilla",
    "ES70": "Canarias",
    "PT11": "Norte (PT)", "PT15": "Algarve (PT)", "PT16": "Centro (PT)",
    "PT17": "Lisboa (PT)", "PT18": "Alentejo (PT)", "PT20": "Azores (PT)", "PT30": "Madeira (PT)",
}


# ------------------------------------------------- derivaciones
def ccaa_de(d):
    nuts = str(d.get("nuts") or "")
    for pref, nombre in NUTS2_CCAA.items():
        if nuts.startswith(pref):
            return nombre
    if d.get("region"):
        return str(d["region"])
    return "Portugal" if d.get("pais") == "Portugal" else "Sin territorio"


def territorio_de(d):
    return str(d.get("territorio") or "") or ccaa_de(d)


def servicio_de(d):
    t = normalizar(" ".join([str(d.get("titulo", "")), " ".join(d.get("senales", {}).get("sin_zanja", []))]))
    cpvs = " ".join(str(c) for c in d.get("cpv", []))
    if d.get("senales", {}).get("sin_zanja"):
        return "Rehabilitación sin zanja"
    if d.get("es_redaccion_proyecto") or cpvs.startswith("713") or " 713" in " " + cpvs:
        return "Redacción de proyecto e ingeniería"
    if any(p in t for p in ("conservacion", "mantenimiento", "manutencao", "conservacao")):
        return "Conservación y mantenimiento de redes"
    if any(p in t for p in ("edar", "etar", "depuradora", "depuracion", "tratamiento de aguas")):
        return "Depuración y EDAR"
    if any(p in t for p in ("renovacion", "renovacao", "sustitucion", "rehabilitacion", "reabilitacao")):
        return "Renovación y obra de redes"
    return "Otros (ciclo del agua)"


def dias_plazo(d, hoy):
    plazo = str(d.get("plazo_presentacion", ""))[:10]
    if len(plazo) != 10:
        return None
    try:
        return (datetime.strptime(plazo, "%Y-%m-%d").date() - hoy).days
    except ValueError:
        return None


def urgencia_de(d, hoy):
    dias = dias_plazo(d, hoy)
    if dias is None:
        return "SIN PLAZO PUBLICADO", None
    if dias < 0:
        return "VENCIDA", dias
    if dias <= 7:
        return "CRÍTICA (esta semana)", dias
    if dias <= 21:
        return "ALTA (menos de 3 semanas)", dias
    if dias <= 45:
        return "MEDIA", dias
    return "AMPLIA", dias


def anio_de(d):
    for campo in ("fecha_adjudicacion", "fecha_deteccion"):
        v = str(d.get(campo, ""))[:4]
        if v.isdigit():
            return int(v)
    return ahora_utc().year


def es_cerrada(d):
    return str(d.get("estado", "")).strip().upper() in ("ADJ", "RES", "ANUL")


def vigente(d, hoy):
    if es_cerrada(d):
        return False
    dias = dias_plazo(d, hoy)
    return dias is None or dias >= 0


# ------------------------------------------------- piezas de maquetación
def _marca_seccion(ws, fila, col, titulo, subtitulo=None):
    ws.cell(fila, col).fill = FILL_ROJO
    ws.column_dimensions[get_column_letter(col)].width = 2.2
    c = ws.cell(fila, col + 1, titulo)
    c.font = F_TITULO
    if subtitulo:
        s = ws.cell(fila + 1, col + 1, subtitulo)
        s.font = F_SUB
    return fila + (3 if subtitulo else 2)


def _cabeceras(ws, fila, cols, anchos):
    for i, (nombre, ancho) in enumerate(zip(cols, anchos), 1):
        c = ws.cell(fila, i, nombre)
        c.font = F_HDR
        c.fill = FILL_HUESO
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = LINEA
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[fila].height = 26
    ws.freeze_panes = ws.cell(fila + 1, 1)
    return fila + 1


def _celda(ws, fila, col, valor, fuente=F_TXT, fmt=None, alin=WRAP):
    c = ws.cell(fila, col, valor)
    c.font = fuente
    c.alignment = alin
    c.border = LINEA
    if fmt:
        c.number_format = fmt
    return c


def _enlace(ws, fila, col, url):
    if not url:
        return
    c = _celda(ws, fila, col, "Ver expediente", Font(name=FN, size=10, bold=True, color=ROJO))
    c.hyperlink = url


# ------------------------------------------------- hojas
def _hoja_portada(wb, historico, hoy, urls):
    ws = wb.create_sheet("Portada")
    ws.sheet_view.showGridLines = False
    for col, w in (("A", 2.2), ("B", 34), ("C", 26), ("D", 26), ("E", 26), ("F", 10)):
        ws.column_dimensions[col].width = w
    ws.cell(2, 2, "canalis").font = Font(name="Rockwell Light", size=18, color="0C2638")
    c = ws.cell(2, 5, "RADAR DE LICITACIONES  ES / PT")
    c.font = Font(name=FN, size=8, bold=True, color=GRIS)
    c.alignment = DER
    fila = _marca_seccion(ws, 5, 1, "Excel Histórico del Radar",
                          f"Repositorio automático. Se regenera en cada barrido. Última actualización: {ahora_utc().strftime('%d/%m/%Y %H:%M')} UTC")
    vivos = [d for d in historico.values() if vigente(d, hoy)]
    adjs = [d for d in historico.values() if d.get("adjudicatarios") or es_cerrada(d)]
    kpis = [
        ("Detecciones acumuladas", len(historico)),
        ("Vigentes ahora mismo", len(vivos)),
        ("Importe vigente (EUR)", round(sum(d.get("importe_eur") or 0 for d in vivos))),
        ("Adjudicaciones registradas", len(adjs)),
    ]
    fila += 1
    for i, (lbl, val) in enumerate(kpis):
        col = 2 + i
        c = ws.cell(fila, col, lbl); c.font = F_KPI_LBL; c.fill = FILL_HUESO
        c.alignment = Alignment(wrap_text=True, vertical="top")
        n = ws.cell(fila + 1, col, val); n.font = F_KPI_NUM; n.fill = FILL_HUESO
        n.alignment = Alignment(vertical="bottom")
        if "EUR" in lbl:
            n.number_format = '#,##0'
        ws.column_dimensions[get_column_letter(col)].width = 26
    ws.row_dimensions[fila].height = 30
    ws.row_dimensions[fila + 1].height = 52
    fila += 4
    fila = _marca_seccion(ws, fila, 1, "Cómo usar este archivo", None)
    guia = [
        ("Vigentes", "Las oportunidades vivas, ordenadas por urgencia. Lo que vence antes, arriba. Usa los filtros de la cabecera (triángulo de cada columna) para acotar por comunidad, servicio o importe."),
        ("Adjudicaciones", "Quién ha ganado qué. Filtra por la columna Ganador para ver todo lo que ha ganado una empresa; por Año o CCAA para acotar. Ejemplo: Ganador = Insituform y Año = 2026."),
        ("Ranking por año", "Las 15 empresas que más licitaciones de nuestro ámbito ganan cada año (Ranking 2025, Ranking 2026...), con contratos e importe."),
        ("Análisis", "Resumen por comunidad autónoma y por tipo de servicio: cuántas, cuánto dinero, cuántas vivas."),
        ("Histórico por año", "El archivo completo, una pestaña por año. Nada se borra."),
    ]
    for nombre, texto in guia:
        ws.cell(fila, 2, nombre).font = F_NEGRITA
        c = ws.cell(fila, 3, texto); c.font = F_TXT_GRIS; c.alignment = WRAP
        ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=5)
        ws.row_dimensions[fila].height = 30
        fila += 1
    fila += 1
    c = ws.cell(fila, 2, "EXCEL INTERNO (ficha comercial del equipo)")
    c.font = Font(name=FN, size=10, bold=True, color=ROJO)
    c.hyperlink = urls["interno"]
    c2 = ws.cell(fila + 1, 2, "Panel web del radar")
    c2.font = Font(name=FN, size=10, bold=True, color=ROJO)
    c2.hyperlink = urls["panel"]
    aviso = ws.cell(fila + 3, 2, "Verifica siempre las fechas e importes en la fuente oficial antes de actuar y traslada las oportunidades maduras al Excel interno")
    aviso.font = Font(name=FN, size=9, color=GRISCLARO)
    ws.merge_cells(start_row=fila + 3, start_column=2, end_row=fila + 3, end_column=5)


COLS_VIG = ["Urgencia", "Días de plazo", "Score", "Entidad / órgano", "CCAA", "Territorio",
            "Servicio", "Objeto", "Importe (EUR)", "Plazo", "Señales sin zanja", "Detectado", "Enlace"]
ANCH_VIG = [22, 9, 7, 26, 16, 14, 24, 52, 14, 11, 20, 11, 12]


def _hoja_vigentes(wb, historico, hoy):
    ws = wb.create_sheet("Vigentes")
    ws.sheet_view.showGridLines = False
    fila = _marca_seccion(ws, 1, 1, "Oportunidades vigentes",
                          "Ordenadas por urgencia del plazo y puntuación. Filtra con los triángulos de la cabecera.")
    vivos = [d for d in historico.values() if vigente(d, hoy)]
    def clave(d):
        dias = dias_plazo(d, hoy)
        return (dias if dias is not None else 9999, -(d.get("score") or 0))
    fila = _cabeceras(ws, fila, COLS_VIG, ANCH_VIG)
    ini = fila
    for d in sorted(vivos, key=clave):
        urg, dias = urgencia_de(d, hoy)
        critica = dias is not None and dias <= 7
        _celda(ws, fila, 1, urg, F_ROJO if critica else (F_NEGRITA if dias is not None and dias <= 21 else F_TXT_GRIS))
        _celda(ws, fila, 2, dias, F_TXT, "0", CENTRO)
        _celda(ws, fila, 3, d.get("score"), F_ROJO if (d.get("score") or 0) >= 75 else F_TXT, "0", CENTRO)
        _celda(ws, fila, 4, d.get("entidad_censo") or d.get("organo", ""), F_NEGRITA)
        _celda(ws, fila, 5, ccaa_de(d))
        _celda(ws, fila, 6, territorio_de(d))
        _celda(ws, fila, 7, servicio_de(d))
        _celda(ws, fila, 8, (d.get("titulo") or "")[:250])
        _celda(ws, fila, 9, d.get("importe_eur"), F_TXT, EUR, DER)
        _celda(ws, fila, 10, str(d.get("plazo_presentacion", ""))[:10])
        _celda(ws, fila, 11, ", ".join(d.get("senales", {}).get("sin_zanja", [])))
        _celda(ws, fila, 12, d.get("fecha_deteccion", ""))
        _enlace(ws, fila, 13, d.get("enlace"))
        fila += 1
    ws.auto_filter.ref = f"A{ini - 1}:M{max(fila - 1, ini)}"


COLS_ADJ = ["Año", "Fecha adjudicación", "Ganador", "Entidad / órgano", "CCAA", "Servicio",
            "Objeto", "Importe adjudicación (EUR)", "Importe licitación (EUR)", "Baja %", "Ofertas", "Enlace"]
ANCH_ADJ = [7, 12, 34, 26, 16, 24, 50, 15, 15, 8, 8, 12]


def _hoja_adjudicaciones(wb, historico, hoy):
    ws = wb.create_sheet("Adjudicaciones")
    ws.sheet_view.showGridLines = False
    fila = _marca_seccion(ws, 1, 1, "Adjudicaciones",
                          "Quién ha ganado qué. Filtra por Ganador, Año o CCAA para responder cualquier pregunta.")
    adjs = [d for d in historico.values() if d.get("adjudicatarios") or es_cerrada(d)]
    fila = _cabeceras(ws, fila, COLS_ADJ, ANCH_ADJ)
    ini = fila
    for d in sorted(adjs, key=lambda x: (str(x.get("fecha_adjudicacion") or x.get("fecha_deteccion") or "")), reverse=True):
        ganadores = d.get("adjudicatarios") or []
        imp_a, imp_l = d.get("importe_adjudicacion"), d.get("importe_eur")
        baja = round((1 - imp_a / imp_l) * 100, 1) if imp_a and imp_l and imp_l > 0 else None
        _celda(ws, fila, 1, anio_de(d), F_TXT, "0", CENTRO)
        _celda(ws, fila, 2, d.get("fecha_adjudicacion") or "")
        _celda(ws, fila, 3, " + ".join(ganadores) if ganadores else "No consta en la publicación",
               F_NEGRITA if ganadores else F_TXT_GRIS)
        _celda(ws, fila, 4, d.get("entidad_censo") or d.get("organo", ""))
        _celda(ws, fila, 5, ccaa_de(d))
        _celda(ws, fila, 6, servicio_de(d))
        _celda(ws, fila, 7, (d.get("titulo") or "")[:250])
        _celda(ws, fila, 8, imp_a, F_NEGRITA, EUR, DER)
        _celda(ws, fila, 9, imp_l, F_TXT, EUR, DER)
        _celda(ws, fila, 10, baja, F_TXT, '0.0"%"', CENTRO)
        _celda(ws, fila, 11, d.get("num_ofertas"), F_TXT, "0", CENTRO)
        _enlace(ws, fila, 12, d.get("enlace"))
        fila += 1
    ws.auto_filter.ref = f"A{ini - 1}:L{max(fila - 1, ini)}"


def _hoja_ranking(wb, historico, anio):
    ws = wb.create_sheet(f"Ranking {anio}")
    ws.sheet_view.showGridLines = False
    fila = _marca_seccion(ws, 1, 1, f"Ranking de ganadores {anio}",
                          "Las 15 empresas con más adjudicaciones del año en nuestro ámbito. Importe como desempate. Reparto por igual en las UTE multi-lote.")
    stats = defaultdict(lambda: {"n": 0, "eur": 0.0, "ccaa": defaultdict(int), "ultima": ""})
    for d in historico.values():
        ganadores = d.get("adjudicatarios") or []
        if not ganadores or anio_de(d) != anio:
            continue
        parte = (d.get("importe_adjudicacion") or 0) / len(ganadores)
        for g in ganadores:
            e = stats[g]
            e["n"] += 1
            e["eur"] += parte
            e["ccaa"][ccaa_de(d)] += 1
            f = str(d.get("fecha_adjudicacion") or d.get("fecha_deteccion") or "")
            e["ultima"] = max(e["ultima"], f)
    cols = ["#", "Empresa", "Adjudicaciones", "Importe total (EUR)", "Importe medio (EUR)", "CCAA principal", "Última adjudicación"]
    fila = _cabeceras(ws, fila, cols, [5, 42, 13, 17, 16, 18, 14])
    orden = sorted(stats.items(), key=lambda kv: (-kv[1]["n"], -kv[1]["eur"]))[:15]
    for i, (empresa, e) in enumerate(orden, 1):
        _celda(ws, fila, 1, i, F_ROJO if i <= 3 else F_TXT, "0", CENTRO)
        _celda(ws, fila, 2, empresa, F_NEGRITA)
        _celda(ws, fila, 3, e["n"], F_TXT, "0", CENTRO)
        _celda(ws, fila, 4, round(e["eur"]), F_NEGRITA, EUR, DER)
        _celda(ws, fila, 5, round(e["eur"] / e["n"]) if e["n"] else None, F_TXT, EUR, DER)
        _celda(ws, fila, 6, max(e["ccaa"].items(), key=lambda kv: kv[1])[0] if e["ccaa"] else "")
        _celda(ws, fila, 7, e["ultima"][:10])
        fila += 1
    if not orden:
        _celda(ws, fila, 2, "Aún sin adjudicaciones con ganador registrado este año. Se llena sola con los barridos diarios.", F_TXT_GRIS)


def _tabla_resumen(ws, fila, titulo, clave_fn, historico, hoy):
    fila = _marca_seccion(ws, fila, 1, titulo, None)
    grupos = defaultdict(lambda: {"n": 0, "vig": 0, "adj": 0, "eur_v": 0.0, "eur_a": 0.0})
    for d in historico.values():
        g = grupos[clave_fn(d)]
        g["n"] += 1
        if vigente(d, hoy):
            g["vig"] += 1
            g["eur_v"] += d.get("importe_eur") or 0
        if d.get("adjudicatarios") or es_cerrada(d):
            g["adj"] += 1
            g["eur_a"] += d.get("importe_adjudicacion") or 0
    cols = ["Grupo", "Detecciones", "Vigentes", "Adjudicadas", "Importe vigente (EUR)", "Importe adjudicado (EUR)"]
    fila = _cabeceras(ws, fila, cols, [30, 11, 10, 11, 18, 19])
    for nombre, g in sorted(grupos.items(), key=lambda kv: -(kv[1]["eur_v"] + kv[1]["eur_a"])):
        _celda(ws, fila, 1, nombre, F_NEGRITA)
        _celda(ws, fila, 2, g["n"], F_TXT, "0", CENTRO)
        _celda(ws, fila, 3, g["vig"], F_TXT, "0", CENTRO)
        _celda(ws, fila, 4, g["adj"], F_TXT, "0", CENTRO)
        _celda(ws, fila, 5, round(g["eur_v"]), F_TXT, EUR, DER)
        _celda(ws, fila, 6, round(g["eur_a"]), F_TXT, EUR, DER)
        fila += 1
    return fila + 2


def _hoja_analisis(wb, historico, hoy):
    ws = wb.create_sheet("Análisis")
    ws.sheet_view.showGridLines = False
    fila = _tabla_resumen(ws, 1, "Por comunidad autónoma", ccaa_de, historico, hoy)
    ws.auto_filter.ref = None
    _tabla_resumen(ws, fila, "Por servicio", servicio_de, historico, hoy)


COLS_HIST = ["Detectado", "Estado", "Score", "Entidad / órgano", "CCAA", "Servicio", "Objeto",
             "Importe (EUR)", "Ganador", "Importe adjudicación (EUR)", "Enlace"]
ANCH_HIST = [11, 9, 7, 26, 16, 24, 52, 14, 30, 15, 12]


def _hojas_historico(wb, historico):
    por_anio = defaultdict(list)
    for d in historico.values():
        por_anio[anio_de(d)].append(d)
    for anio in sorted(por_anio, reverse=True):
        ws = wb.create_sheet(f"Histórico {anio}")
        ws.sheet_view.showGridLines = False
        fila = _marca_seccion(ws, 1, 1, f"Histórico {anio}",
                              "Archivo completo del año. Nada se borra: las cerradas quedan como memoria del mercado.")
        fila = _cabeceras(ws, fila, COLS_HIST, ANCH_HIST)
        ini = fila
        for d in sorted(por_anio[anio], key=lambda x: str(x.get("fecha_deteccion", "")), reverse=True):
            ganadores = d.get("adjudicatarios") or []
            _celda(ws, fila, 1, d.get("fecha_deteccion", ""))
            _celda(ws, fila, 2, d.get("estado", ""), F_TXT, None, CENTRO)
            _celda(ws, fila, 3, d.get("score"), F_TXT, "0", CENTRO)
            _celda(ws, fila, 4, d.get("entidad_censo") or d.get("organo", ""))
            _celda(ws, fila, 5, ccaa_de(d))
            _celda(ws, fila, 6, servicio_de(d))
            _celda(ws, fila, 7, (d.get("titulo") or "")[:250])
            _celda(ws, fila, 8, d.get("importe_eur"), F_TXT, EUR, DER)
            _celda(ws, fila, 9, " + ".join(ganadores))
            _celda(ws, fila, 10, d.get("importe_adjudicacion"), F_TXT, EUR, DER)
            _enlace(ws, fila, 11, d.get("enlace"))
            fila += 1
        ws.auto_filter.ref = f"A{ini - 1}:K{max(fila - 1, ini)}"


def generar(historico, urls):
    hoy = ahora_utc().date()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_portada(wb, historico, hoy, urls)
    _hoja_vigentes(wb, historico, hoy)
    _hoja_adjudicaciones(wb, historico, hoy)
    anios = sorted({anio_de(d) for d in historico.values()} | {hoy.year}, reverse=True)
    for anio in anios:
        _hoja_ranking(wb, historico, anio)
    _hoja_analisis(wb, historico, hoy)
    _hojas_historico(wb, historico)
    DOCS.mkdir(exist_ok=True)
    wb.save(DOCS / "Radar_Historico_ES-PT.xlsx")
